"""
Excel row matching and cell writing for the SPS notification tracking workbook.
Finds the pre-populated row by 문서번호 and fills in all computed/LLM fields.
"""
import io
import re
import shutil
import zipfile
from datetime import date, datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from lxml import etree as _ET

# OpenXML namespace URIs used throughout the xlsx XML
_SS  = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
_REL = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
_PKG = 'http://schemas.openxmlformats.org/package/2006/relationships'

# ── Column index mapping (1-based) ────────────────────────────────────────────
COL = {
    '담당자':        1,
    '순번':          2,
    '중요도':        3,
    '통보유형':      4,
    '통보국':        5,
    '배포일':        6,
    '문서번호':      7,
    '제목':          8,
    '내용':          9,
    '해당품목':     10,
    '목적':         11,
    '해당국가':     12,
    '의견마감일':   13,
    '발효일':       14,
    '국내수출품목': 15,
    '관련부서':     16,
    '주간보고':     17,
    '구분':         18,
    '품목':         19,
    '검토메모':     20,
}

# openpyxl fill objects kept for the legacy write_fields path
YELLOW_FILL = PatternFill('solid', fgColor='FFFF00')
LIME_FILL   = PatternFill('solid', fgColor='CCFF99')
NO_FILL     = PatternFill('none')
_LIME_RGBS  = frozenset({'FFCCFF99', '00CCFF99', 'CCFF99'})

# Fields the tool writes (skips pre-filled identification fields)
WRITABLE_FIELDS = [
    '중요도', '통보유형', '통보국', '제목', '내용', '해당품목', '목적', '해당국가',
    '의견마감일', '발효일', '국내수출품목', '관련부서', '주간보고', '구분', '품목',
]

# Always overwritten even if the cell already has a value
FORCE_WRITE_FIELDS = {'통보국', '통보유형'}


# ── Shared low-level helpers ──────────────────────────────────────────────────

def _has_korean(text: str) -> bool:
    return bool(re.search(r'[가-힣ᄀ-ᇿ㄰-㆏]', text))


def _normalize_doc_number(doc_num: str) -> str:
    return re.sub(r'\s+', '', doc_num).upper()


def _col_num(ref: str) -> int:
    """'AB5' → 28  (1-based column index from the letter part of a cell ref)."""
    n = 0
    for ch in ref:
        if ch.isalpha():
            n = n * 26 + (ord(ch.upper()) - 64)
        else:
            break
    return n


def _excel_serial_to_date(value) -> date | None:
    """Convert an Excel date serial (int / float / numeric string) to a Python date."""
    try:
        # Excel epoch is 1899-12-30, accounting for Excel's spurious 1900 leap-year bug
        return date(1899, 12, 30) + timedelta(days=int(float(value)))
    except (ValueError, TypeError, OverflowError):
        return None


# ── XML cell helpers (used by the direct-patch path) ─────────────────────────

def _cell_xml_value(cell_el, ss_strings: list) -> str:
    """Return the plain-text value stored in a worksheet <c> element."""
    t = cell_el.get('t', '')
    v = cell_el.find(f'{{{_SS}}}v')
    if t == 's' and v is not None:
        try:
            return ss_strings[int(v.text or '')]
        except (ValueError, IndexError):
            return ''
    if t == 'inlineStr':
        return ''.join(el.text or '' for el in cell_el.iter(f'{{{_SS}}}t'))
    return (v.text or '') if v is not None else ''


def _is_lime_xml(cell_el, xfs: list, fills: list) -> bool:
    """Return True if this cell element has a lime (CCFF99) solid fill."""
    s_attr = cell_el.get('s')
    if s_attr is None:
        return False
    try:
        xf = xfs[int(s_attr)]
    except (IndexError, ValueError):
        return False
    fill_id = xf.get('fillId')
    if fill_id is None:
        return False
    try:
        fill = fills[int(fill_id)]
    except (IndexError, ValueError):
        return False
    pf = fill.find(f'{{{_SS}}}patternFill')
    if pf is None or pf.get('patternType') != 'solid':
        return False
    fg = pf.find(f'{{{_SS}}}fgColor')
    return fg is not None and 'CCFF99' in fg.get('rgb', '').upper()


def _get_or_add_fill_xf(styles_root, color_hex: str,
                         xfs_cache: list, fills_cache: list) -> int:
    """
    Find or add a solid-fill XF entry in the parsed styles.xml element tree.
    Returns the index into <cellXfs> for the cell's 's' attribute.
    Only appends — never modifies existing entries, so existing style indices
    remain valid for all other cells.
    """
    fills_el = styles_root.find(f'{{{_SS}}}fills')
    xfs_el   = styles_root.find(f'{{{_SS}}}cellXfs')
    if fills_el is None or xfs_el is None:
        return 0

    # 1 — find existing solid fill with this colour
    fill_idx = None
    for i, fill in enumerate(fills_cache):
        pf = fill.find(f'{{{_SS}}}patternFill')
        if pf is not None and pf.get('patternType') == 'solid':
            fg = pf.find(f'{{{_SS}}}fgColor')
            if fg is not None and color_hex.upper() in fg.get('rgb', '').upper():
                fill_idx = i
                break

    if fill_idx is None:
        new_fill = _ET.SubElement(fills_el, f'{{{_SS}}}fill')
        pf = _ET.SubElement(new_fill, f'{{{_SS}}}patternFill')
        pf.set('patternType', 'solid')
        fg = _ET.SubElement(pf, f'{{{_SS}}}fgColor')
        fg.set('rgb', f'FF{color_hex.upper()}')
        _ET.SubElement(pf, f'{{{_SS}}}bgColor').set('indexed', '64')
        fills_cache.append(new_fill)
        fill_idx = len(fills_cache) - 1
        fills_el.set('count', str(fill_idx + 1))

    # 2 — find existing XF using this fill with applyFill=1
    for i, xf in enumerate(xfs_cache):
        if xf.get('fillId') == str(fill_idx) and xf.get('applyFill') in ('1', 'true'):
            return i

    # 3 — add a new XF (copy base attributes from XF 0, then override fill)
    base_attrs = dict(xfs_cache[0].attrib) if xfs_cache else {}
    new_xf = _ET.SubElement(xfs_el, f'{{{_SS}}}xf')
    for k, v in base_attrs.items():
        new_xf.set(k, v)
    new_xf.set('fillId',    str(fill_idx))
    new_xf.set('applyFill', '1')
    new_xf.set('numFmtId',  new_xf.get('numFmtId', '0'))
    new_xf.set('fontId',    new_xf.get('fontId',   '0'))
    new_xf.set('borderId',  new_xf.get('borderId', '0'))
    xfs_cache.append(new_xf)
    xf_idx = len(xfs_cache) - 1
    xfs_el.set('count', str(xf_idx + 1))
    return xf_idx


# ── ZIP / XML internals ───────────────────────────────────────────────────────

def _styles_insert_fill_xf(styles_bytes: bytes, color_hex: str) -> tuple:
    """
    Find or append a solid-fill XF entry in styles.xml using string surgery.
    Returns (new_bytes, xf_index).

    Preserves ALL existing XML byte-for-byte — only inserts text before closing
    tags.  This avoids lxml's round-trip serialization artefacts (namespace-prefix
    changes, attribute reordering) that corrupt Excel's theme-color font references
    in cells that were never touched by this tool.
    """
    text        = styles_bytes.decode('utf-8')
    rgb_upper   = color_hex.upper()          # e.g. 'CCFF99'
    rgb_argb    = 'FF' + rgb_upper           # e.g. 'FFCCFF99'

    # ── Find existing fill with this colour ──────────────────────────────────
    fills_start = text.find('<fills')
    fills_end   = text.find('</fills>') + len('</fills>')

    fill_idx  = None
    if fills_start != -1 and fills_end > fills_start:
        fills_block = text[fills_start:fills_end]
        fill_starts = [m.start() for m in re.finditer(r'<fill\b', fills_block)]
        for i, pos in enumerate(fill_starts):
            close = fills_block.find('</fill>', pos)
            segment = fills_block[pos: close + len('</fill>')] if close != -1 else fills_block[pos:]
            if 'solid' in segment and (rgb_upper in segment.upper() or rgb_argb in segment.upper()):
                fill_idx = i
                break

    fill_count_m = re.search(r'<fills\b[^>]*\bcount="(\d+)"', text)
    n_fills = int(fill_count_m.group(1)) if fill_count_m else len(fill_starts) if fills_start != -1 else 0

    if fill_idx is None:
        fill_idx = n_fills
        new_fill_xml = (
            f'<fill><patternFill patternType="solid">'
            f'<fgColor rgb="{rgb_argb}"/>'
            f'<bgColor indexed="64"/>'
            f'</patternFill></fill>'
        )
        text = text.replace('</fills>', new_fill_xml + '</fills>', 1)
        if fill_count_m:
            text = re.sub(
                r'(<fills\b[^>]*\bcount=")(\d+)(")',
                lambda m: m.group(1) + str(n_fills + 1) + m.group(3),
                text, count=1,
            )

    # ── Find existing XF that uses this fill ────────────────────────────────
    xfs_start = text.find('<cellXfs')
    xfs_end   = text.find('</cellXfs>') + len('</cellXfs>')

    xf_idx = None
    if xfs_start != -1 and xfs_end > xfs_start:
        xfs_block   = text[xfs_start:xfs_end]
        xf_starts   = [m.start() for m in re.finditer(r'<xf\b', xfs_block)]
        for i, pos in enumerate(xf_starts):
            sc = xfs_block.find('/>', pos)
            segment = xfs_block[pos: sc + 2] if sc != -1 else xfs_block[pos: pos + 300]
            if f'fillId="{fill_idx}"' in segment and 'applyFill="1"' in segment:
                xf_idx = i
                break

    xf_count_m = re.search(r'<cellXfs\b[^>]*\bcount="(\d+)"', text)
    n_xfs = int(xf_count_m.group(1)) if xf_count_m else (len(xf_starts) if xfs_start != -1 else 0)

    if xf_idx is None:
        xf_idx = n_xfs
        new_xf_xml = (
            f'<xf numFmtId="0" fontId="0" fillId="{fill_idx}" '
            f'borderId="0" xfId="0" applyFill="1"/>'
        )
        text = text.replace('</cellXfs>', new_xf_xml + '</cellXfs>', 1)
        if xf_count_m:
            text = re.sub(
                r'(<cellXfs\b[^>]*\bcount=")(\d+)(")',
                lambda m: m.group(1) + str(n_xfs + 1) + m.group(3),
                text, count=1,
            )

    return text.encode('utf-8'), xf_idx


def _zip_load(excel_path: str) -> tuple:
    """Read every file from the xlsx ZIP. Returns (raw_dict, infos_dict)."""
    with zipfile.ZipFile(excel_path, 'r') as z:
        infos = {info.filename: info for info in z.infolist()}
        raw   = {name: z.read(name) for name in z.namelist()}
    return raw, infos


def _zip_save(excel_path: str, raw: dict, infos: dict) -> None:
    """Repack the xlsx ZIP, preserving each file's original compression type."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w') as out_z:
        for name, content in raw.items():
            compress = infos[name].compress_type if name in infos else zipfile.ZIP_DEFLATED
            out_z.writestr(name, content, compress_type=compress)
    with open(excel_path, 'wb') as f:
        f.write(buf.getvalue())


def _xml_get_sheet_file(raw: dict, target_month: str = None) -> tuple:
    """
    Locate the target month's worksheet file inside the xlsx ZIP.
    Returns (sheet_file_path, sheet_name).
    """
    wb_root   = _ET.fromstring(raw['xl/workbook.xml'])
    rels_root = _ET.fromstring(raw['xl/_rels/workbook.xml.rels'])

    sheets = wb_root.findall(f'{{{_SS}}}sheets/{{{_SS}}}sheet')

    def _pick_rid(sheets_list):
        # Priority 1: target_month
        if target_month:
            for sh in sheets_list:
                if target_month in sh.get('name', ''):
                    return sh.get(f'{{{_REL}}}id'), sh.get('name')
        # Priority 2: current year/month
        now = datetime.now()
        month_str = f'{str(now.year)[2:]}.{now.month}월'
        for sh in sheets_list:
            if month_str in sh.get('name', ''):
                return sh.get(f'{{{_REL}}}id'), sh.get('name')
        # Priority 3: any sheet with '월' that isn't the manual
        for sh in sheets_list:
            name = sh.get('name', '')
            if '월' in name and '매뉴얼' not in name:
                return sh.get(f'{{{_REL}}}id'), name
        return None, None

    rid, sheet_name = _pick_rid(sheets)
    if rid is None:
        raise ValueError('적합한 월별 시트를 찾을 수 없습니다.')

    for rel in rels_root.findall(f'{{{_PKG}}}Relationship'):
        if rel.get('Id') == rid:
            tgt = rel.get('Target')
            sheet_file = f'xl/{tgt}' if not tgt.startswith('/') else tgt.lstrip('/')
            if sheet_file in raw:
                return sheet_file, sheet_name

    raise ValueError(f'워크시트 파일을 찾을 수 없습니다 (rId={rid})')


def _xml_load_strings(raw: dict) -> list:
    """Parse xl/sharedStrings.xml into a plain list of strings."""
    if 'xl/sharedStrings.xml' not in raw:
        return []
    ss_root = _ET.fromstring(raw['xl/sharedStrings.xml'])
    return [
        ''.join(t.text or '' for t in si.iter(f'{{{_SS}}}t'))
        for si in ss_root.findall(f'{{{_SS}}}si')
    ]


def _xml_find_row(ws_root, ss_strings: list, doc_number: str) -> tuple:
    """
    Parse the worksheet XML to:
      1. Build col_map from the header row (row 1).
      2. Find the data row whose 문서번호 cell matches doc_number.

    Returns (row_idx, col_map, base_date_raw_str, row_element)
    or raises ValueError if not found.
    """
    sd = ws_root.find(f'{{{_SS}}}sheetData')
    if sd is None:
        raise ValueError('sheetData not found in worksheet XML')

    rows = sd.findall(f'{{{_SS}}}row')

    # Build col_map from header row
    col_map = dict(COL)
    for row_el in rows:
        if int(row_el.get('r', 0)) != 1:
            continue
        detected = {}
        for c_el in row_el.findall(f'{{{_SS}}}c'):
            val = _cell_xml_value(c_el, ss_strings).strip()
            if val in COL:
                detected[val] = _col_num(c_el.get('r', ''))
        if len(detected) >= len(COL) // 2:
            col_map = {**COL, **detected}
        break

    doc_col  = col_map.get('문서번호', COL['문서번호'])
    date_col = col_map.get('배포일',   COL['배포일'])
    needle      = _normalize_doc_number(doc_number)
    needle_base = re.sub(r'/ADD\.\d+$', '', needle)

    for row_el in rows:
        r_num = int(row_el.get('r', 0))
        if r_num < 2:
            continue

        # Build a quick ref→element map for this row
        cell_map = {_col_num(c.get('r', '')): c for c in row_el.findall(f'{{{_SS}}}c')}

        doc_el = cell_map.get(doc_col)
        if doc_el is None:
            continue

        raw_val = _cell_xml_value(doc_el, ss_strings)
        if not raw_val:
            continue

        cell_val = _normalize_doc_number(raw_val)
        cell_ids = [_normalize_doc_number(x) for x in re.split(r'[,;]', cell_val)]
        if not (needle in cell_ids or needle == cell_val or
                (needle_base != needle and needle_base in cell_ids)):
            continue

        # Found — extract base_date raw value
        date_el = cell_map.get(date_col)
        base_date_raw = _cell_xml_value(date_el, ss_strings) if date_el is not None else ''
        return r_num, col_map, base_date_raw, row_el

    raise ValueError(f'문서번호 {doc_number}을(를) Excel에서 찾을 수 없습니다.')


# ── Main write path ───────────────────────────────────────────────────────────

def _direct_patch_xlsx(
    excel_path: str,
    doc_number: str,
    target_month: str,
    fields: dict,
    uncertain_fields: list,
    is_non_english: bool,
) -> tuple:
    """
    Single-pass direct XML patch: locate the target row and write cells
    without ever letting openpyxl rebuild styles.xml.

    Non-target cells are never parsed as writable objects, so their font
    colours, fills, and every other format attribute survive exactly as-is.

    Returns (row_idx, base_date).
    """
    raw, infos = _zip_load(excel_path)

    sheet_file, _sheet_name = _xml_get_sheet_file(raw, target_month)
    ss_strings = _xml_load_strings(raw)

    ws_root = _ET.fromstring(raw[sheet_file])
    row_idx, col_map, base_date_raw, row_el = _xml_find_row(ws_root, ss_strings, doc_number)

    # Resolve base_date for return (caller may use it for date calculations)
    from date_engine import parse_excel_date
    base_date = (
        _excel_serial_to_date(base_date_raw)
        or parse_excel_date(base_date_raw)
        if base_date_raw else None
    )

    # Parse styles once for fill lookup and lime-cell detection
    styles_root = _ET.fromstring(raw['xl/styles.xml'])
    xfs_list   = list(styles_root.findall(f'{{{_SS}}}cellXfs/{{{_SS}}}xf'))
    fills_list  = list(styles_root.findall(f'{{{_SS}}}fills/{{{_SS}}}fill'))

    # Build cell map for target row
    cell_map = {_col_num(c.get('r', '')): c for c in row_el.findall(f'{{{_SS}}}c')}

    force_write = FORCE_WRITE_FIELDS | ({'해당품목', '목적', '발효일'} if is_non_english else set())

    writes = []  # (col_idx, str_value, fill_hex | None)
    for field_name in WRITABLE_FIELDS:
        value = fields.get(field_name)
        if value is None:
            continue
        col_idx = col_map.get(field_name)
        if col_idx is None:
            continue

        c_el = cell_map.get(col_idx)

        if c_el is not None and field_name not in force_write:
            cur = _cell_xml_value(c_el, ss_strings)
            if cur and cur.strip():
                # Preserve unless: lime-fill cell with non-Korean content (portal pre-fill)
                if _has_korean(cur) or not _is_lime_xml(c_el, xfs_list, fills_list):
                    continue

        fill_hex = None
        if field_name in uncertain_fields:
            fill_hex = 'FFFF00'
        elif is_non_english and field_name in ('제목', '내용'):
            fill_hex = 'CCFF99'

        writes.append((col_idx, str(value), fill_hex))

    # Reviewer notes
    reportable = [f for f in uncertain_fields if col_map.get(f)]
    if reportable:
        memo_col = col_map.get('검토메모', COL['검토메모'])
        memo_el  = cell_map.get(memo_col)
        if not (memo_el is not None and _cell_xml_value(memo_el, ss_strings).strip()):
            writes.append((memo_col, '검토 필요: ' + ', '.join(reportable), None))

    if writes:
        # Resolve fill XF indices via byte surgery — never re-serialise styles.xml
        # through lxml, which changes namespace prefixes and corrupts theme-colour
        # font references in every cell that uses those XF entries.
        fill_xf: dict = {}
        for _, _, fill_hex in writes:
            if fill_hex and fill_hex not in fill_xf:
                new_styles, xf_idx = _styles_insert_fill_xf(raw['xl/styles.xml'], fill_hex)
                fill_xf[fill_hex] = xf_idx
                raw['xl/styles.xml'] = new_styles  # accumulate changes across colours

        # Apply writes
        for col_idx, value, fill_hex in writes:
            c_el = cell_map.get(col_idx)
            if c_el is None:
                c_el = _ET.Element(f'{{{_SS}}}c')
                c_el.set('r', f'{get_column_letter(col_idx)}{row_idx}')
                cell_map[col_idx] = c_el

            if fill_hex and fill_hex in fill_xf:
                c_el.set('s', str(fill_xf[fill_hex]))
            # else: keep original 's' (style index) untouched

            for tag in (f'{{{_SS}}}v', f'{{{_SS}}}is', f'{{{_SS}}}f'):
                for el in list(c_el.findall(tag)):
                    c_el.remove(el)
            c_el.attrib.pop('t', None)
            c_el.set('t', 'inlineStr')
            is_el = _ET.SubElement(c_el, f'{{{_SS}}}is')
            t_el  = _ET.SubElement(is_el, f'{{{_SS}}}t')
            t_el.text = value
            if value.startswith(' ') or value.endswith(' '):
                t_el.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')

        # Rebuild row with cells in left-to-right order (required by Excel)
        for c in list(row_el.findall(f'{{{_SS}}}c')):
            row_el.remove(c)
        for c in sorted(cell_map.values(), key=lambda c: _col_num(c.get('r', 'A'))):
            row_el.append(c)

        raw[sheet_file] = _ET.tostring(
            ws_root, xml_declaration=True, encoding='UTF-8', standalone=True
        )
        _zip_save(excel_path, raw, infos)

    return row_idx, base_date


# ── Public API ────────────────────────────────────────────────────────────────

def load_and_process(excel_path: str, doc_number: str, fields: dict,
                     uncertain_fields: list, is_non_english: bool = False,
                     target_month: str = None) -> tuple:
    """
    Locate the pre-populated row for doc_number and write computed fields.
    Returns (success: bool, error_msg: str, row_idx: int | None).
    """
    try:
        shutil.copy2(excel_path, excel_path + '.sps_bak')
        row_idx, _ = _direct_patch_xlsx(
            excel_path, doc_number, target_month,
            fields, uncertain_fields, is_non_english,
        )
        return True, '', row_idx
    except PermissionError:
        return False, 'Excel 파일이 다른 프로그램에서 열려 있습니다. 닫고 다시 시도해주세요.', None
    except Exception as e:
        return False, str(e), None


def get_base_date(excel_path: str, doc_number: str, target_month: str = None):
    """
    Return the 배포일 date for a given document number.
    Uses direct XML reading — no openpyxl overhead.
    """
    try:
        raw, _ = _zip_load(excel_path)
        sheet_file, _ = _xml_get_sheet_file(raw, target_month)
        ss_strings = _xml_load_strings(raw)
        ws_root = _ET.fromstring(raw[sheet_file])
        _, _, base_date_raw, _ = _xml_find_row(ws_root, ss_strings, doc_number)
        if not base_date_raw:
            return None
        from date_engine import parse_excel_date
        return _excel_serial_to_date(base_date_raw) or parse_excel_date(base_date_raw)
    except Exception:
        return None


# ── Legacy / compat (not called by the main path) ─────────────────────────────

def _is_lime_cell(cell) -> bool:
    try:
        fill = cell.fill
        return fill.patternType == 'solid' and fill.fgColor.rgb.upper() in _LIME_RGBS
    except Exception:
        return False


def _detect_col_map(ws) -> dict:
    detected = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        name = str(cell.value).strip()
        if name in COL:
            detected[name] = cell.column
    if len(detected) >= len(COL) // 2:
        return {**COL, **detected}
    return dict(COL)


def _get_month_sheet(wb, target_month: str = None):
    if target_month:
        for name in wb.sheetnames:
            if target_month in name:
                return wb[name]
    now = datetime.now()
    month_str = f'{str(now.year)[2:]}.{now.month}월'
    for name in wb.sheetnames:
        if month_str in name:
            return wb[name]
    for name in wb.sheetnames:
        if '매뉴얼' not in name and '월' in name:
            return wb[name]
    return None


def find_row(wb, doc_number: str, target_month: str = None):
    """Legacy: find row via openpyxl (used by get_base_date compat callers)."""
    ws = _get_month_sheet(wb, target_month)
    if ws is None:
        return None, None, None, dict(COL)
    col_map = _detect_col_map(ws)
    needle = _normalize_doc_number(doc_number)
    needle_base = re.sub(r'/ADD\.\d+$', '', needle)
    doc_col = col_map['문서번호']
    date_col = col_map['배포일']
    for row in ws.iter_rows(min_row=2):
        cell = row[doc_col - 1]
        if cell.value is None:
            continue
        cell_val = _normalize_doc_number(str(cell.value))
        cell_ids = [_normalize_doc_number(x) for x in re.split(r'[,;]', cell_val)]
        if needle in cell_ids or needle == cell_val or (needle_base != needle and needle_base in cell_ids):
            base_date = None
            date_cell = row[date_col - 1]
            if date_cell.value:
                from date_engine import parse_excel_date
                base_date = parse_excel_date(date_cell.value)
            return ws, cell.row, base_date, col_map
    return None, None, None, col_map


def _find_row_readonly(excel_path: str, doc_number: str, target_month: str = None):
    """Legacy: read-only openpyxl row search (superseded by _xml_find_row)."""
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        ws = _get_month_sheet(wb, target_month)
        if ws is None:
            return None, None, None, dict(COL)
        sheet_name = ws.title
        col_map = _detect_col_map(ws)
        needle = _normalize_doc_number(doc_number)
        needle_base = re.sub(r'/ADD\.\d+$', '', needle)
        doc_col = col_map['문서번호']
        date_col = col_map['배포일']
        for row in ws.iter_rows(min_row=2):
            cell = row[doc_col - 1]
            if cell.value is None:
                continue
            cell_val = _normalize_doc_number(str(cell.value))
            cell_ids = [_normalize_doc_number(x) for x in re.split(r'[,;]', cell_val)]
            if needle in cell_ids or needle == cell_val or (needle_base != needle and needle_base in cell_ids):
                base_date = None
                date_cell = row[date_col - 1]
                if date_cell.value:
                    from date_engine import parse_excel_date
                    base_date = parse_excel_date(date_cell.value)
                return sheet_name, cell.row, base_date, col_map
    finally:
        wb.close()
    return None, None, None, dict(COL)


def write_fields(ws, row_idx: int, fields: dict, uncertain_fields: list,
                 is_non_english: bool = False, col_map: dict = None):
    """Legacy: openpyxl-based cell writer (superseded by _direct_patch_xlsx)."""
    if col_map is None:
        col_map = COL
    force_write = FORCE_WRITE_FIELDS | ({'해당품목', '목적', '발효일'} if is_non_english else set())
    for field_name in WRITABLE_FIELDS:
        if field_name not in fields:
            continue
        col_idx = col_map.get(field_name)
        if col_idx is None:
            continue
        cell = ws.cell(row=row_idx, column=col_idx)
        if field_name not in force_write and cell.value not in (None, ''):
            if _is_lime_cell(cell) and not _has_korean(str(cell.value)):
                pass
            else:
                continue
        value = fields[field_name]
        if value is None:
            continue
        cell.value = value
        if field_name in uncertain_fields:
            cell.fill = YELLOW_FILL
        elif is_non_english and field_name in ('제목', '내용'):
            cell.fill = LIME_FILL
    reportable_flags = [f for f in uncertain_fields if col_map.get(f)]
    if reportable_flags:
        memo_col = col_map.get('검토메모', COL.get('검토메모'))
        if memo_col:
            note_cell = ws.cell(row=row_idx, column=memo_col)
            if note_cell.value in (None, ''):
                note_cell.value = '검토 필요: ' + ', '.join(reportable_flags)
    return True
